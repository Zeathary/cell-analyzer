'''
@brief Helper File containing functions to export recorded data about cells to either an excel spreadsheet or a csv file
'''

import os
import csv
import openpyxl
import math

'''
    Exports given cell data to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param coordinates: dictionary indexed by cell id containing coordinates 
    @param areas: dictionary indexed by cell id containing areas
    @param headers: iterable object containing headers for each column
'''
def culture_to_excel_file(filename, coordinates, areas, time_between_frames, area_of_frame, coordinate_headers=None, area_headers=None, units="mm"):
    coordinates_to_excel_file(filename, coordinates, coordinate_headers, "Positions")
    area_to_excel_file(filename, areas, area_headers, "Areas")
    # TODO Implement Separate stats function for image files
    stats = calc_culture_cell_statistics(coordinates, areas, time_between_frames, area_of_frame, units=units)
    culture_stats_to_excel_file(filename, stats, "Culture Stats")


'''
    Exports given dictionary containing data on an individuals cell's area and coordinates to excel file
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary containing data about the cell
    @param headers: iterable object containing headers for each column
'''
def individual_to_excel_file(filename, data:dict, time_between_frames, units="mm", sheetname=None):
    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
        # Create new sheet to use for this data
        sheet = wb.create_sheet(sheetname)

    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()
        # Rename Default sheet and set it as our active one
        sheet = wb.active
        sheet.title = sheetname

    # Write Data to sheet
    # Loop through dictionary. For every key write all of its values within the same column then move onto the next
    for key, value in data.items():
        current_row = 1
        # Add Header
        sheet.cell(current_row, current_col, key)
        current_row += 1

        # Add Data
        for entry in value:
            sheet.cell(current_row, current_col, float(entry))
            current_row += 1

        current_col += 1

    # Generate Stats only if the given data was a video and therefore had multiple frames
    keys = list(data.keys())
    if len(data[keys[0]]) > 1:
        # Generate Statistics using positional and area data about cells and export that to same excel sheet
        coordinates = merge(data[f"X Position ({units})"], data[f"Y Position ({units})"])
        stats = calc_individual_cell_statistics(coordinates, data[f'Area ({units}^2)'], time_between_frames, units=units)

        # Keep Track of the column stats start on
        stats_col = current_col
        # Reset Row
        current_row = 1

        # Add Stats Headers
        sheet.cell(current_row, current_col, "Statistic")
        current_col += 1
        sheet.cell(current_row, current_col, "Value")
        current_col += 1
        current_row += 1

        # Loop Through Stats and add them to excel sheet
        for key, value in stats.items():
            # Add Header
            current_col = stats_col
            sheet.cell(current_row, current_col, key)
            current_col += 1

            # Add Data
            sheet.cell(current_row, current_col, value)
            current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Exports given cell coordinates to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary indexed by cell id, and containing tuples of the coordinates for each cell
    @param headers: iterable object containing headers for each column
'''
def coordinates_to_excel_file(filename, data, headers=None, sheetname=None):
    if headers is None:
        headers = []

    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
        # Add Sheet to store column/row data about this iteration
        sheet = wb.create_sheet(sheetname)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()
        # Rename default sheet created and continue
        sheet = wb.active
        sheet.title = sheetname



    # Write Data to sheet
    # Create Headers
    for i in range(0, len(headers)):
        # Cell (row, col, data) Base 1
        sheet.cell(current_row, current_col + i, headers[i])
    current_row += 1

    # Loop through all data given then extract useful info and append it
    # Adds data to new row. Argument must be iterable object
    for key, value in data.items():
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1
        for val in value:
            # Split tuples into x and y coordinate
            val = tuple(val)
            x = val[0]
            y = val[1]

            # Place x coordinate in one column and y in the next
            sheet.cell(current_row, current_col, float(x))
            current_col += 1
            sheet.cell(current_row, current_col, float(y))
            current_col += 1

        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Exports given cell areas to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param data: Dictionary indexed by cell id, and containing data about each cell
    @param headers: iterable object containing headers for each column
'''
def area_to_excel_file(filename, data, headers=None, sheetname=None):
    if headers is None:
        headers = []

    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()

    # Add Sheet to store column/row data about this iteration
    sheet = wb.create_sheet(sheetname)

    # Write Data to sheet
    # Create Headers
    for i in range(0, len(headers)):
        # Cell (row, col, data) Base 1
        sheet.cell(current_row, current_col + i, headers[i])
    current_row += 1

    # Loop through all data given then extract useful info and append it
    # Adds data to new row. Argument must be iterable object
    for key, value in data.items():
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1
        for val in value:
            # Remove [] from string to format it correctly for Excel
            val = str(val).replace("[", "")
            val = str(val).replace("]", "")
            # Convert Area to float to avoid number stored as text
            sheet.cell(current_row, current_col, float(val))
            current_col += 1

        # Insert Excel Formula to display Total Growth the cell Underwent
        #=INDEX(B142:DR142,MATCH(TRUE,INDEX((B142:DR142<>0),0),0))
        growth_formula = f'=INDIRECT(ADDRESS({current_row}, {current_col - 1})) - INDEX(INDIRECT(ADDRESS({current_row}, 2)):INDIRECT(ADDRESS({current_row}, {current_col})),MATCH(TRUE,INDEX((INDIRECT(ADDRESS({current_row}, 2)):INDIRECT(ADDRESS({current_row}, {current_col}))<>0),0),0))'
        sheet.cell(current_row, current_col, growth_formula)
        current_col += 1
        # Insert Excel Formula to display largest amount of change between two time intervals
        change_formula = f'=_xlfn.AGGREGATE(14, 6, INDIRECT(ADDRESS({current_row}, 2)):INDIRECT(ADDRESS({current_row}, {current_col - 2}))-INDIRECT(ADDRESS({current_row}, 3)):INDIRECT(ADDRESS({current_row}, {current_col - 1})), 1)'
        sheet.cell(current_row, current_col, change_formula)
        current_col += 1
        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Calculates Statistics from positional data from one cell. These Stats include:
    Total Displacement: Total distance moved, Final Distance from origin, maximum distance from origin, average distance from origin,
    Maximum speed, average speed, average angle of direction, and final angle of direction between final point and origin
    @param data List of tuples/list containing x/y coordinates of given cells location
    @param area a list of all recorded areas the cell had each frame
    @param time_between_frames Time in minutes between each frame of cell growth video
    @return Dictionary containing statistics generated 
'''
def calc_individual_cell_statistics(data, areas, time_between_frames, units="mm"):
    # Create dictionary to hold all calculated statistics
    stats = {}
    distances = []
    origin_distances = []
    # In mm / min
    speeds = []
    # Angle in degrees between last and current point
    angle_of_direction = []
    final_angle = 0

    if data is not None:
        # Grab origin point
        origin_x = data[0][0]
        origin_y = data[0][1]
        x = 0
        y = 0

        # Loop through all positions and calculate stats between points
        for i in range(1, len(data)):
            # Grab x and y coordinates
            x = data[i][0]
            y = data[i][1]
            prevx = data[i-1][0]
            prevy = data[i-1][1]

            # Calc distance from traveled between each step
            distance = math.dist([prevx, prevy], [x, y])
            distances.append(distance)
            # Calc distance between origin and current step
            origin_distances.append(math.dist([origin_x, origin_y], [x, y]))
            # calc current speed
            speeds.append(distance/time_between_frames)
            # calc angle of direction from last point. Because of the way opencv stores coordinates
            # ((0,0) would be the top left) we need to convert the angle by subtracting 360 degrees by it
            angle = 360 - ((math.atan2(y - prevy, x - prevx) * (180 / math.pi)) % 360)
            angle_of_direction.append(angle)

            # If on final coordinate calculate angle between this and origin point
            if i == (len(data) - 1):
                final_angle = 360 - ((math.atan2(y - origin_y, x - origin_x) * (180 / math.pi)) % 360)

        # Positional Stats
        # Total Displacement (Total Distance Traveled)
        stats[f"Total Displacement ({units})"] = sum(distances)
        # Final Distace from Origin
        stats[f"Final Distance from Origin ({units})"] = math.dist([origin_x, origin_y], [x, y])
        # Maximum Distance from origin
        stats[f"Maximum Distance from Origin ({units})"] = max(origin_distances)
        # Average Distance from origin
        stats[f"Average Distance from Origin ({units})"] = sum(origin_distances)/len(origin_distances)
        # Maximum Distance Traveled in one Interval
        stats[f"Maximum Distance Traveled in one Interval ({units})"] = max(distances)
        # Max Speed (distance/time)
        stats[f"Maximum Speed ({units}/min)"] = max(speeds)
        # Average Speed
        stats[f"Average Speed ({units}/min)"] = sum(speeds)/len(speeds)
        # Average Angle of direction from origin in degrees
        stats[f"Average Angle of Direction from Origin (degrees)"] = sum(angle_of_direction)/len(angle_of_direction)
        # Angle of direction from origin to final point
        stats["Angle of Direction between Origin and Final Point (degrees)"] = final_angle
        # Categorize Direction of Movement
        compass_brackets = ["E", "NE", "N", "NW", "W", "SW", "S", "SE", "E"]
        compass_lookup = round(final_angle / 45)
        stats["Compass Direction Moved"] = compass_brackets[compass_lookup]

        # Area Stats
        # Max Size
        stats[f"Maximum Size ({units}^2)"] = max(areas)
        # Min Size
        stats[f"Minimum Size ({units}^2)"] = min(areas)
        # Average Size
        stats[f"Average Size ({units}^2)"] = sum(areas)/len(areas)
        # Calc change in size of the cell between first and last frame
        stats[f"Change in Cell Size ({units}^2)"] = areas[len(areas) - 1] - areas[0]
        # Calc Average Change in growth between each time interval
        change = 0
        for i in range(1, len(areas)):
            change += areas[i] - areas[i-1]
        avg_change = change/len(areas)
        stats[f"Average Change in Cell Size Between one Interval ({units}^2)"] = avg_change
    else:
        raise Exception("Empy Data Set Given")

    return stats


'''
    Exports given cell areas to an excel spreadsheet
    @param filename: Name of excel file to edit or write to. Should end with extension .xls or .xlsx
    @param stats: Dictionary of stats gathered about a culture. Keys will be used as headers
    @param sheetname Name of the created sheet in excel document
'''
def culture_stats_to_excel_file(filename, stats, sheetname=None):
    current_row = 1
    current_col = 1

    # If filename does not end in .xls extension exit
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise Exception("File must be of type .xls or .xlsx")

    # If file already exists, create a new sheet to store data on
    if os.path.exists(f"{filename}"):
        # Open excel file for reading and writing
        wb = openpyxl.load_workbook(filename)
    else:
        # Otherwise create a new excel file to write to
        # Create Workbook to store all sheets and their data
        wb = openpyxl.Workbook()

    # Add Sheet to store column/row data about this iteration
    sheet = wb.create_sheet(sheetname)

    # Write Headers
    sheet.cell(current_row, current_col, "Statistic")
    current_col += 1
    sheet.cell(current_row, current_col, "Value")
    current_col += 1
    current_row += 1

    # Loop Through Stats and add them to excel sheet
    for key, value in stats.items():
        # Add Header
        current_col = 1
        sheet.cell(current_row, current_col, key)
        current_col += 1

        # Add Data
        sheet.cell(current_row, current_col, value)
        current_row += 1

    # Save File
    wb.save(f"{filename}")


'''
    Calculates Statistics averaging out data from all cells
    Total Displacement: Total distance moved, Final Distance from origin, maximum distance from origin, average distance from origin,
    Maximum speed, average speed, average angle of direction, and final angle of direction between final point and origin
    @param data Dictionary indexed by cell id containing a list of tuples/list containing x/y coordinates of given cells location
    @param time_between_frames Time in minutes between each frame of cell growth video
    @return Dictionary containing statistics generated 
'''
def calc_culture_cell_statistics(positional_data, area_data, time_between_frames, area_of_frame, units="mm"):
    # Create dictionary to hold all calculated statistics
    stats = {}
    displacements = []
    final_distances = []
    # In mm / min
    speeds = []
    # Angle in degrees between last and current point
    angle_of_direction = []
    final_sizes = []
    growth = []
    max_cell_size = 0
    max_cell_id = 0
    min_cell_size = None
    min_cell_id = 0


    for key, data in positional_data.items():
        if data is not None:
            data = tuple(data)
            # Grab origin point
            origin_x = data[0][0]
            origin_y = data[0][1]
            x = 0
            y = 0
            distances = []

            # Loop through all positions and calculate stats between points
            for i in range(1, len(data)):
                # Only do stats on non zero values as zero is simply a placeholder specifying that the cell was not tracked that frame
                if data[i][0] != 0 or data[i][1] != 0:
                    # Grab x and y coordinates
                    x = data[i][0]
                    y = data[i][1]
                    prevx = data[i-1][0]
                    prevy = data[i-1][1]


                    # Calc Distance traveled between frames
                    distance = math.dist([prevx, prevy], [x, y])
                    distances.append(distance)

                    # calc current speed
                    speeds.append(distance/time_between_frames)

                    # If on final coordinate calculate total displacement, angle and distance between this and origin point
                    # Because of the way opencv stores coordinates
                    # ((0,0) would be the top left) we need to convert the angle by subtracting 360 degrees by it
                    if i == (len(data) - 1):
                        final_angle = 360 - ((math.atan2(y - origin_y, x - origin_x) * (180 / math.pi)) % 360)
                        angle_of_direction.append(final_angle)

                        # Calc final distance from origin
                        final_distances.append(math.dist([origin_x, origin_y], [x, y]))

                        # Record total displacement
                        displacements.append(sum(distances))

    # Generate Stats based on Cell Size
    for key, value in area_data.items():
        value = list(value)

        # Check if this cell is the new largest or smallest cell in the culture
        if min(value) != 0 and (min_cell_size is None or min(value) < min_cell_size):
            min_cell_size = min(value)
            min_cell_id = key
        if max_cell_size < max(value):
            max_cell_size = max(value)
            max_cell_id = key

        # Record Final Size for the cell
        final_sizes.append(value[len(value)-1])
        # Find first non zero value within its list of areas and record the index
        start_index = value.index(next(filter(lambda x: x!=0, value)))
        # Record difference between the initial cell and its final size
        growth.append(value[len(value) - 1] - value[start_index])



    # TODO FIX Division by zero
    # If the data analyzed was a video and therefore there were multiple frames to compare against then calc stats based on movement/change
    if len(displacements) != 0:
        # Total Displacement (distance traveled throughout whole video)
        stats[f"Average Total Displacement ({units})"] = sum(displacements)/len(displacements)
        stats[f"Max Distance Traveled by one Cell ({units})"] = max(displacements)
        stats[f"Min Distance Traveled by one Cell ({units})"] = min(displacements)
        # Average Distance from origin
        stats[f"Average Final Distance from Origin ({units})"] = sum(final_distances)/len(final_distances)
        # Average Speed
        stats[f"Average Speed ({units}/min)"] = sum(speeds)/len(speeds)
        # Maximum Recorded Speed
        stats[f"Maximum Recorded Speed ({units}/min)"] = max(speeds)
        # Minimum Recorded Speed
        stats[f"Minimum Recorded Speed ({units}/min)"] = min(speeds)
        # Angle of direction from origin to final point
        stats[f"Average Angle of Direction between Origin and Final Point (degrees)"] = sum(angle_of_direction)/len(angle_of_direction)
        # Categorize Direction of Movement
        compass_brackets = ["E", "NE", "N", "NW", "W", "SW", "S", "SE", "E"]
        compass_lookup = round(stats["Average Angle of Direction between Origin and Final Point (degrees)"] / 45)
        stats["Average Compass Direction Moved"] = compass_brackets[compass_lookup]
        # Average cell growth/shrinkage
        stats[f"Average Change in Cell Size ({units}^2)"] = sum(growth) / len(growth)

    # Calculate All Stats that can be done on both images and video
    # Calculate Final Frame's Confluency
    # Percentage of Frame the cells take up
    stats["Final Frame's Confluency (%)"] = sum(final_sizes) / area_of_frame
    # Largest Recorded Cell
    stats[f"Largest Cell ({units}^2)"] = max_cell_size
    stats[f"Largest Cell's ID"] = max_cell_id
    # Smallest Recorded Cell
    stats[f"Smallest Cell ({units}^2)"] = min_cell_size
    stats["Smallest Cell's ID"] = min_cell_id
    # Average Size of Cells
    stats[f"Average Final Size of Cell ({units}^2)"] = sum(final_sizes) / len(final_sizes)


    return stats


'''
    Exports given cell data to a comma separated value file. Does not calculate statistics
    @param filename: Name of File to append to or save data to. Should end with extension .csv and contain full path as necessary
           If csv file already exists data will be appended to the end without new headers
    @param data: Dictionary containing mapping to iterable objects each containing data about one aspect of the cell. ie data["X Position"] = []
'''
def individual_to_csv_file(filename, data:dict):
    # If filename does not end in .xls extension exit
    if not filename.endswith(".csv"):
        raise Exception("File must be of type .csv")

    # Grab all keys to the data dictionary given
    headers = list(data.keys())
    # record the number of data points in each entry in the dict(aka the number of frames)
    num_data_points = len(data[headers[0]])

    # If file already exists, append data to the end without adding new headers
    if os.path.exists(f"{filename}"):
        # Open csv file in append mode
        with open(filename, "a", newline='') as file:
            csvwriter = csv.writer(file)

            # Write Data file one row at a time
            # Loop through each frame of data. For every frame assemble a row of data to append to the file
            for i in range(0, num_data_points):
                row = list()
                # Loop through each list in data and append the item at the current data point/frame to our row list
                for key in headers:
                    row.append(data[key][i])

                # Append data to new row in csv file
                csvwriter.writerow(row)

    else:
        # Otherwise create a new csv file to write to
        with open(filename, "w", newline='') as file:
            csvwriter = csv.writer(file)

            # Write headers (keys to dict)
            csvwriter.writerow(headers)

            # Write Data file one row at a time
            # Loop through each frame of data. For every frame assemble a row of data to append to the file
            for i in range(0, num_data_points):
                row = list()
                # Loop through each list in data and append the item at the current data point/frame to our row list
                for key in headers:
                    row.append(data[key][i])

                # Append data to new row in csv file
                csvwriter.writerow(row)


'''
    Exports given raw cell data to a comma separated value file. Does not calculate statistics
    @param filename: Name of File to append to or save data to. Should end with extension .csv and contain full path as necessary
           If csv file already exists data will be appended to the end without new headers
    @param positional data: Dictionary containing mapping between a cell id and an iterable object of tuples (x,y coordinates)
    @param area data: Dictionary containing mapping between a cell id and an iterable object of areas(floating point numbers)
    @param positional_headers: Iterable object of headers for each column of the csv file
    @param positional_headers: Iterable object of headers for each column of the csv file
'''
def culture_to_csv_file(filename, positional_data:dict, area_data:dict, positional_headers=None, area_headers=None):
    if area_headers is None:
        area_headers = []
    if positional_headers is None:
        positional_headers = []

    # If filename does not end in .xls extension exit
    if not filename.endswith(".csv"):
        raise Exception("File must be of type .csv")
    elif os.path.exists(filename):
        raise Exception("File already exists")

    # Grab all cell ids / keys to the two dictionaries
    ids = list(positional_data.keys())
    # record the number of data points in each entry in the dict(aka the number of frames)
    num_data_points = len(positional_data[ids[0]])


    # Otherwise create a new csv file to write to
    with open(filename, "w", newline='') as file:
        csvwriter = csv.writer(file)

        # Write headers (keys to dict)
        csvwriter.writerow(positional_headers + area_headers)

        # Write Data file one row at a time
        # Loop through each cell's data and create one unified list and append it to the next available row
        for i in range(0, len(ids)):
            # Add cell id to list to start
            row = list()
            row.append(ids[i])

            # Loop through all positional data for this cell and append it to the list
            for coordinate in positional_data[ids[i]]:
                x = coordinate[0]
                y = coordinate[1]
                row.append(x)
                row.append(y)

            # Loop through all size data for this cell and append it to the list
            for size in area_data[ids[i]]:
                row.append(size)

            # Append data to new row in csv file
            csvwriter.writerow(row)


'''
    Combines two lists together into one list containing tuples of (list1[i], list2[i])
    @param list1 List of elements the same length as list2
    @param list2 List of elements the same length as list1
    @return Merged List of tuples
'''
def merge(list1, list2):
    merged_list = [(list1[i], list2[i]) for i in range(0, len(list1))]
    return merged_list